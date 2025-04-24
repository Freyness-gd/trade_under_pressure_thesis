import time
from io import BytesIO

import requests
from openpyxl import load_workbook

# --- Config ---
reporter_codes = [
  "ALB",
  "DZA",
  "AND",
  "AGO",
  "ATG",
  "ARG",
  "AUS",
  "AUT",
  "BHS",
  "BHR",
  "BGD",
  "BRB",
  "BEL",
  "BLZ",
  "BEN",
  "BTN",
  "BOL",
  "BWA",
  "BRA",
  "BRN",
  "BGR",
  "BFA",
  "BDI",
  "CPV",
  "KHM",
  "CMR",
  "CAN",
  "CAF",
  "TCD",
  "CHL",
  "CHN",
  "COL",
  "COM",
  "COD",
  "CRI",
  "CIV",
  "CUB",
  "CYP",
  "DNK",
  "DMA",
  "DOM",
  "ECU",
  "EGY",
  "SLV",
  "GNQ",
  "SWZ",
  "ETH",
  "FRO",
  "FJI",
  "FIN",
  "FRA",
  "GAB",
  "GMB",
  "DEU",
  "GHA",
  "GRC",
  "GRL",
  "GRD",
  "GTM",
  "GIN",
  "GNB",
  "GUY",
  "HTI",
  "HND",
  "HKG",
  "HUN",
  "ISL",
  "IND",
  "IDN",
  "IRN",
  "IRQ",
  "IRL",
  "ISR",
  "ITA",
  "JAM",
  "JPN",
  "JOR",
  "KEN",
  "KIR",
  "KOR",
  "KWT",
  "LSO",
  "LBR",
  "LBY",
  "LUX",
  "MDG",
  "MWI",
  "MYS",
  "MDV",
  "MLI",
  "MLT",
  "MHL",
  "MRT",
  "MUS",
  "MEX",
  "FSM",
  "MNG",
  "MAR",
  "MMR",
  "NAM",
  "NRU",
  "NPL",
  "NLD",
  "NZL",
  "NIC",
  "NER",
  "NGA",
  "NOR",
  "OMN",
  "PAK",
  "PLW",
  "PAN",
  "PNG",
  "PRY",
  "PER",
  "PHL",
  "PRT",
  "PRI",
  "QAT",
  "RWA",
  "WSM",
  "STP",
  "SAU",
  "SEN",
  "SYC",
  "SLE",
  "SGP",
  "SLB",
  "SOM",
  "ZAF",
  "ESP",
  "LKA",
  "KNA",
  "LCA",
  "VCT",
  "SDN",
  "SUR",
  "SWE",
  "CHE",
  "SYR",
  "TZA",
  "THA",
  "TGO",
  "TON",
  "TTO",
  "TUN",
  "TUR",
  "TUV",
  "UGA",
  "ARE",
  "GBR",
  "USA",
  "URY",
  "VUT",
  "ZMB",
  "ZWE",
]
# reporter_codes = ["ABCD"]
base_url = (
  "https://wits.worldbank.org/Download.aspx?"
  "StartYear=1988&EndYear=2023&Tradeflow=Import"
  "&Indicator=MPRT-TRD-VL&Partner=ALL&Type=PartnerTimeseries&Lang=en"
)

# --- Download and Modify ---
for code in reporter_codes:
  url = f"{base_url}&Reporter={code}"
  filename = f"IMPORT_{code}.xlsx"
  print(f"Downloading: {filename} ...")

  try:
    response = requests.get(url)
    if response.status_code == 200:
      # Load workbook from memory
      workbook = load_workbook(BytesIO(response.content))
      sheet_to_remove = workbook.sheetnames[0]  # Always remove the first sheet
      workbook.remove(workbook[sheet_to_remove])
      print(f"Removed '{sheet_to_remove}' from {filename}")

      workbook.save(filename)
      print(f"Saved: {filename}")
    else:
      print(f"Failed to download {code}: HTTP {response.status_code}")
  except Exception as e:
    print(f"Error processing {code}: {e}")

  time.sleep(0.5)  # Wait 500ms between downloads
